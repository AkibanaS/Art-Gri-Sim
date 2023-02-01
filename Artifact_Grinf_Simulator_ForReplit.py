# -*- coding: utf-8 -*-
from PIL import ImageDraw, ImageFont, ImageFilter, ImageFile, Image
from openpyxl.styles import Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl import load_workbook
import openpyxl
import os
import io
import time
import random
import discord
from flask import Flask
import string
import asyncio

app = Flask(__name__)


@app.route('/')
def index():
    return 'Hello Heroku_Flask'


if __name__ == '__main__':
    app.run()

# This example requires the 'message_content' intent.

# import部

os.chdir(os.path.dirname(os.path.abspath(__file__)))

# 定義部
TOKEN = 'DISCORD_BOT_TOKEN'

intents = discord.Intents.default()
intents.message_content = True

client = discord.Client(intents=intents)

# 聖遺物の名前
# 1.モンド地域
Artifact_name0 = ['剣闘士のフィナーレ', '大地を流浪する楽団']
Artifact_name1 = ['雷のような怒り', '雷を鎮める尊者']
Artifact_name2 = ['翠緑の影', '愛される少女']
Artifact_name3 = ['氷風を彷徨う勇士', '沈淪の心']
# 2.璃月地域
Artifact_name4 = ['燃え盛る炎の魔女', '烈火を渡る賢者']
Artifact_name5 = ['旧貴族のしつけ', '血染めの騎士道']
Artifact_name6 = ['悠久の磐岩', '逆飛びの流星']
Artifact_name7 = ['千岩牢固', '蒼白の炎']
Artifact_name8 = ['辰砂往生録', '来歆の余響']
# 3.稲妻地域
Artifact_name9 = ['追憶のしめ縄', '絶縁の旗印']
Artifact_name10 = ['華館夢醒形骸記', '海染硨磲']
# 4.スメール地域
Artifact_name11 = ['森林の記憶', '金メッキの夢']
Artifact_name12 = ['砂上の楼閣の史話', '楽園の絶花']
# 5.後日追加分
Artifact_nameN = ['', '']


# 聖遺物部位の名前
Artifact_kinds = ['生の花', '死の羽', '時の砂', '空の杯', '理の冠']
Artifact_kinds_image = ['Flower', 'Plume', 'Sands', 'Goblet', 'Circlet']
# 1.モンド地域
Artifact_kind0 = ['剣闘士の未練', '剣闘士の帰着', '剣闘士の希望', '剣闘士の酩酊', '剣闘士の凱旋']
Artifact_kind1 = ['楽団の朝の光', '琴師の矢羽', 'フィナーレの時計', '吟遊者の水筒', '指揮者のハット']
Artifact_kind2 = ['雷鳥の憐み', '雷災の生存者', '雷霆の時計', '落雷の前兆', '雷を呼ぶ冠']
Artifact_kind3 = ['雷討ちの心', '雷討ちの羽根', '雷討ちの刻', '雷討ちの器', '雷討ちの冠']
Artifact_kind4 = ['野花の記憶の草原', '狩人の青緑色の矢羽', '緑の狩人の決心', '緑の狩人の容器', '緑の狩人の冠']
Artifact_kind5 = ['彼方にある少女の心', '少女の揺らぐ思い', '少女の短い華年', '少女の暫く息抜き', '少女の儚き顔']
Artifact_kind6 = ['吹雪の中の思い', '氷を砕く執念', '雪覆う故郷の最後', '霜を纏った気骨', '氷雪を踏む音']
Artifact_kind7 = ['金メッキのコサージュ', '追憶の風', '堅い銅のコンパス', '浮沈の杯', '酒に漬けた帽子']
# 2.璃月地域
Artifact_kind8 = ['魔女の炎の花', '魔女の炎の羽根', '魔女の破滅の時', '魔女の心の炎', '焦げた魔女の帽子']
Artifact_kind9 = ['火渡りの堅実', '火渡りの解放', '火渡りの苦しみ', '火渡りの悟り', '火渡りの知恵']
Artifact_kind10 = ['旧貴族の花', '旧貴族の羽根', '旧貴族の時計', '旧貴族の銀瓶', '旧貴族の仮面']
Artifact_kind11 = ['血染めの鉄の心', '血染めの黒羽', '騎士が血に染めた時', '血染めの騎士のコップ', '血染めの鉄仮面']
Artifact_kind12 = ['盤石芽生の花', '嵯峨連山の翼', '星羅圭玉の日時計', '危岩磐石の杯', '不動玄石の冠']
Artifact_kind13 = ['夏祭りの花', '夏祭りの終わり', '夏祭りの刻', '夏祭りの水風船', '夏祭りの仮面']
Artifact_kind14 = ['偉勲の花', '昭武の羽根', '金銅の日時計', '誓いの金杯', '将師の兜']
Artifact_kind15 = ['無垢の花', '良医の羽', '停頓の時', '超越の杯', '嗤笑の面']
Artifact_kind16 = ['生霊の花', '潜光の羽', '陽轡の遺品', '契約の時', '虺雷の姿']
Artifact_kind17 = ['魂香の花', '垂玉の葉', '祭祀の証', '湧水の杯', '浮流の対玉']
# 3.稲妻地域
Artifact_kind18 = ['羈絆の花', '憶念の矢', '朝露の時', '祈望の心', '無常の面']
Artifact_kind19 = ['威厳の鍔', '切落の羽', '雷雲の印籠', '緋花の壺', '華飾の兜']
Artifact_kind20 = ['栄花の期', '華館の羽', '衆生の歌', '夢醒の瓢箪', '形骸の笠']
Artifact_kind21 = ['海染の花', '淵宮の羽', '別れの貝', '真珠の籠', '海祇の冠']
# 4.スメール地域
Artifact_kind22 = ['迷宮の遊客', '翠蔓の知者', '賢知の定期', '迷いの灯', '月桂の宝冠']
Artifact_kind23 = ['夢境の鉄花', '裁断の羽根', '深金の歳月', '甘露の終宴', '砂王の投影']
Artifact_kind24 = ['諸王の都の始まり', '黄金の邦国の結末', '没落迷途のコア', '迷酔の長夢の守護', '流砂の嗣君の遺宝']
Artifact_kind25 = ['月娘の華彩', '落謝の宴席', '凝結の刹那', '守秘の魔瓶', '紫晶の花冠']

header = ['No.', '聖遺物名', '花', '羽', '砂', '杯', '冠']
header_Article_Main = ['No.', '聖遺物', '部位', '聖遺物名',
                       'Main Op', '値', 'Score', '初期数']
header_Article_Sub = ['HP実数', '攻撃力実数', '防御力実数', 'HP%',
                      '攻撃力%', '防御力%', '会心率', '会心ダメ', '元素熟知', 'チャージ効率']
header_Article = header_Article_Main + header_Article_Sub

# 「Main OPの種類」
# 生の花
Flower_String = 'HP'
Flower_value = '4,780'
# 死の羽
Plume_String = '攻撃力'
Plume_value = '311'
# 時の砂
Sands_String_list = ['攻撃力', '防御力', 'HP', '元素熟知', '元素チャージ効率']
Sands_value_list = ['46.6%', '58.3%', '46.6%', '187', '51.8%']
# 空の杯
Goblet_String_list = ['攻撃力', '防御力', 'HP', '元素熟知', '物理ダメージバフ', '炎元素ダメージバフ',
                      '水元素ダメージバフ', "氷元素ダメージバフ", '雷元素ダメージバフ', '風元素ダメージバフ', '岩元素ダメージバフ', '草元素ダメージバフ']
Goblet_value_list = ['46.6%', '58.3%', '46.6%', '187', '58.3%',
                     '46.6%', '46.6%', '46.6%', '46.6%', '46.6%', '46.6%', '46.6%']
# 理の冠
Circlet_String_list = ['攻撃力', '防御力', 'HP', '元素熟知', '会心率', '会心ダメージ', '与える治癒効果']
Circlet_value_list = ['46.6%', '58.3%',
                      '46.6%', '187', '31.1%', '62.2%', '35.9%']

# 「Sub OPの種類」
SubOP_String = ['HP', '攻撃力', '攻撃力', '防御力',
                'HP', '元素熟知', '元素チャージ効率', '会心率', '会心ダメージ', '防御力']
SubOP_Kind_Num = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9]
SubOP_list_Num = [0, 1, 2, 3, 4, 5, 6, 7, 8]


# SubOPの上昇値
SubOP_0_value = [209.13, 239.00, 269.88, 298.75]
SubOP_1_value = [13.62, 15.56, 17.51, 19.45]
SubOP_2_value = [4.08, 4.66, 5.25, 5.83]
SubOP_3_value = [5.10, 5.83, 6.56, 7.29]
SubOP_4_value = [4.08, 4.66, 5.25, 5.83]
SubOP_5_value = [16.32, 18.65, 20.98, 23.31]
SubOP_6_value = [4.53, 5.18, 5.83, 6.48]
SubOP_7_value = [2.72, 3.11, 3.50, 3.89]
SubOP_8_value = [5.44, 6.22, 6.99, 7.77]
SubOP_9_value = [16.20, 18.52, 20.83, 23.15]

# SubOPの抽選用
SubOP0_Dec_list = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
SubOP1_Dec_list = [1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1]
SubOP2_Dec_list = [2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2]
SubOP3_Dec_list = [3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3]
SubOP4_Dec_list = [4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4]
SubOP5_Dec_list = [5, 5, 5, 5, 5, 5, 5, 5, 5, 5]
SubOP6_Dec_list = [6, 6, 6, 6, 6, 6, 6, 6, 6, 6]
SubOP7_Dec_list = [7, 7, 7, 7, 7, 7, 7, 7]
SubOP8_Dec_list = [8, 8, 8, 8, 8, 8, 8, 8]
SubOP9_Dec_list = [9, 9, 9, 9, 9, 9, 9, 9, 9, 9, 9, 9, 9, 9, 9]


class MainOption_object:
    MainOP_Str_fin = ""
    MainOP_Val_fin = ""
    MainOP_Num_fin = 0


class SubOption_object:
    # SubOPの最終スコア
    SubOP_0_fin = ""
    SubOP_1_fin = ""
    SubOP_2_fin = ""
    SubOP_3_fin = ""

    SubOP_0_fin_value = ""
    SubOP_1_fin_value = ""
    SubOP_2_fin_value = ""
    SubOP_3_fin_value = ""

    SubOP_fir_NUM = ""  # 初期数の表記
    SubOP_fir_n = 0  # 初期数
    SubOP_score_fin = ""
    # SubOPの数値
    SubOP_0 = 0  # 0 HP実数
    SubOP_1 = 0  # 1 攻撃力実数
    SubOP_2 = 0  # 2 攻撃力%
    SubOP_3 = 0  # 3 防御力%
    SubOP_4 = 0  # 4 HP%
    SubOP_5 = 0  # 5 元素熟知
    SubOP_6 = 0  # 6 元素チャージ効率
    SubOP_7 = 0  # 7 会心率
    SubOP_8 = 0  # 8 会心ダメージ
    SubOP_9 = 0  # 9 防御力実数

    # SubOPのスコアカウント
    SubOP_0_count = 0  # 0 HP実数
    SubOP_1_count = 0  # 1 攻撃力実数
    SubOP_2_count = 0  # 2 攻撃力%
    SubOP_3_count = 0  # 3 防御力%
    SubOP_4_count = 0  # 4 HP%
    SubOP_5_count = 0  # 5 元素熟知
    SubOP_6_count = 0  # 6 元素チャージ効率
    SubOP_7_count = 0  # 7 会心率
    SubOP_8_count = 0  # 8 会心ダメージ
    SubOP_9_count = 0  # 9 防御力実数


class Artifact_xlsx_Dec:
    Artifact_Number_0 = ""
    Trial_Number_0 = ""

# 関数部
# MainOP決定


def Artifact_MainOP(MAINOPTION, Kind_number):
    # 花
    if Kind_number == 0:
        MainOP_String = Flower_String    # MainOp決定
        MainOP_Value = Flower_value      # 数値決定
        MainOP_Number = 0                # SubOpで除外
    # 羽
    elif Kind_number == 1:
        MainOP_String = Plume_String    # MainOp決定
        MainOP_Value = Plume_value      # 数値決定
        MainOP_Number = 1               # SubOpで除外
    # 砂
    elif Kind_number == 2:
        MainOP_Dec_Num = random.randrange(
            30)   # 攻撃力,防御力,HP : 26.68%    元素熟知,原チャ効率 : 10%
        # MainOPの抽選
        if 0 <= MainOP_Dec_Num <= 7:
            MainOP_String = Sands_String_list[0]    # MainOp決定
            MainOP_Value = Sands_value_list[0]      # 数値決定
            MainOP_Number = 2                       # SubOpで除外
        elif 8 <= MainOP_Dec_Num <= 15:
            MainOP_String = Sands_String_list[1]    # MainOp決定
            MainOP_Value = Sands_value_list[1]      # 数値決定
            MainOP_Number = 3                       # SubOpで除外
        elif 16 <= MainOP_Dec_Num <= 23:
            MainOP_String = Sands_String_list[2]    # MainOp決定
            MainOP_Value = Sands_value_list[2]      # 数値決定
            MainOP_Number = 4                       # SubOpで除外
        elif 24 <= MainOP_Dec_Num <= 26:
            MainOP_String = Sands_String_list[3]    # MainOp決定
            MainOP_Value = Sands_value_list[3]      # 数値決定
            MainOP_Number = 5                       # SubOpで除外
        elif 27 <= MainOP_Dec_Num <= 29:
            MainOP_String = Sands_String_list[4]    # MainOp決定
            MainOP_Value = Sands_value_list[4]      # 数値決定
            MainOP_Number = 6                       # SubOpで除外
    # 杯
    elif Kind_number == 3:
        # 攻撃力,防御力 : 20%  HP : 17.5%  元素熟知 : 5%  物理ダメージ
        MainOP_Dec_Num = random.randrange(200)
        # MainOPの抽選
        if 0 <= MainOP_Dec_Num <= 39:
            MainOP_String = Goblet_String_list[0]    # MainOp決定
            MainOP_Value = Goblet_value_list[0]      # 数値決定
            MainOP_Number = 2                       # SubOpで除外
        elif 40 <= MainOP_Dec_Num <= 79:
            MainOP_String = Goblet_String_list[1]    # MainOp決定
            MainOP_Value = Goblet_value_list[1]      # 数値決定
            MainOP_Number = 3                       # SubOpで除外
        elif 80 <= MainOP_Dec_Num <= 114:
            MainOP_String = Goblet_String_list[2]    # MainOp決定
            MainOP_Value = Goblet_value_list[2]      # 数値決定
            MainOP_Number = 4                       # SubOpで除外
        elif 115 <= MainOP_Dec_Num <= 119:
            MainOP_String = Goblet_String_list[3]    # MainOp決定
            MainOP_Value = Goblet_value_list[3]      # 数値決定
            MainOP_Number = 5                       # SubOpで除外
        elif 120 <= MainOP_Dec_Num <= 129:
            MainOP_String = Goblet_String_list[4]    # MainOp決定
            MainOP_Value = Goblet_value_list[4]      # 数値決定
            MainOP_Number = 6                       # SubOpで除外
        elif 130 <= MainOP_Dec_Num <= 139:
            MainOP_String = Goblet_String_list[5]    # MainOp決定
            MainOP_Value = Goblet_value_list[5]      # 数値決定
            MainOP_Number = 100                       # SubOpで除外
        elif 140 <= MainOP_Dec_Num <= 149:
            MainOP_String = Goblet_String_list[6]    # MainOp決定
            MainOP_Value = Goblet_value_list[6]      # 数値決定
            MainOP_Number = 100                       # SubOpで除外
        elif 150 <= MainOP_Dec_Num <= 159:
            MainOP_String = Goblet_String_list[7]    # MainOp決定
            MainOP_Value = Goblet_value_list[7]      # 数値決定
            MainOP_Number = 100                       # SubOpで除外
        elif 160 <= MainOP_Dec_Num <= 169:
            MainOP_String = Goblet_String_list[8]    # MainOp決定
            MainOP_Value = Goblet_value_list[8]      # 数値決定
            MainOP_Number = 100                       # SubOpで除外
        elif 170 <= MainOP_Dec_Num <= 179:
            MainOP_String = Goblet_String_list[9]    # MainOp決定
            MainOP_Value = Goblet_value_list[9]      # 数値決定
            MainOP_Number = 100                       # SubOpで除外
        elif 180 <= MainOP_Dec_Num <= 189:
            MainOP_String = Goblet_String_list[10]    # MainOp決定
            MainOP_Value = Goblet_value_list[10]      # 数値決定
            MainOP_Number = 100                        # SubOpで除外
        elif 190 <= MainOP_Dec_Num <= 199:
            MainOP_String = Goblet_String_list[11]    # MainOp決定
            MainOP_Value = Goblet_value_list[11]      # 数値決定
            MainOP_Number = 100                        # SubOpで除外
    # 冠
    elif Kind_number == 4:
        MainOP_Dec_Num = random.randrange(50)
        # MainOPの抽選
        if 0 <= MainOP_Dec_Num <= 10:
            MainOP_String = Circlet_String_list[0]    # MainOp決定
            MainOP_Value = Circlet_value_list[0]      # 数値決定
            MainOP_Number = 2                       # SubOpで除外
        elif 11 <= MainOP_Dec_Num <= 21:
            MainOP_String = Circlet_String_list[1]    # MainOp決定
            MainOP_Value = Circlet_value_list[1]      # 数値決定
            MainOP_Number = 3                       # SubOpで除外
        elif 22 <= MainOP_Dec_Num <= 32:
            MainOP_String = Circlet_String_list[2]    # MainOp決定
            MainOP_Value = Circlet_value_list[2]      # 数値決定
            MainOP_Number = 4                       # SubOpで除外
        elif 33 <= MainOP_Dec_Num <= 34:
            MainOP_String = Circlet_String_list[3]    # MainOp決定
            MainOP_Value = Circlet_value_list[3]      # 数値決定
            MainOP_Number = 5                       # SubOpで除外
        elif 35 <= MainOP_Dec_Num <= 39:
            MainOP_String = Circlet_String_list[4]    # MainOp決定
            MainOP_Value = Circlet_value_list[4]      # 数値決定
            MainOP_Number = 7                      # SubOpで除外
        elif 40 <= MainOP_Dec_Num <= 44:
            MainOP_String = Circlet_String_list[5]    # MainOp決定
            MainOP_Value = Circlet_value_list[5]      # 数値決定
            MainOP_Number = 8                       # SubOpで除外
        elif 45 <= MainOP_Dec_Num <= 49:
            MainOP_String = Circlet_String_list[6]    # MainOp決定
            MainOP_Value = Circlet_value_list[6]      # 数値決定
            MainOP_Number = 100                       # SubOpで除外

    MAINOPTION.MainOP_Str_fin = MainOP_String
    MAINOPTION.MainOP_Val_fin = MainOP_Value
    MAINOPTION.MainOP_Num_fin = MainOP_Number

# SubOP決定


def Artifact_SubOP(SUBOPTION, MainOP_NM):

    SubOP_String = ['HP', '攻撃力', '攻撃力', '防御力',
                    'HP', '元素熟知', '元素チャージ効率', '会心率', '会心ダメージ', '防御力']
    SubOP_Kind_Num = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9]
    SubOP_list_Num = [0, 1, 2, 3, 4, 5, 6, 7, 8]
    # SubOP決定用
    SubOP_decision_list = []
    SubOP_pro_list = []      # 確率
    for i in range(7):
        SubOP_decision_list.append(0)
        SubOP_decision_list.append(9)
        SubOP_decision_list.append(1)
        SubOP_decision_list.append(2)
        SubOP_decision_list.append(3)
        SubOP_decision_list.append(5)
        SubOP_decision_list.append(4)
        SubOP_decision_list.append(1)
        SubOP_decision_list.append(9)
        SubOP_decision_list.append(6)
        SubOP_decision_list.append(7)
        SubOP_decision_list.append(8)
        SubOP_decision_list.append(0)
    for i in range(3):
        SubOP_decision_list.append(2)
        SubOP_decision_list.append(3)
        SubOP_decision_list.append(4)
        SubOP_decision_list.append(5)
        SubOP_decision_list.append(6)
    SubOP_decision_list.append(9)
    SubOP_decision_list.append(8)
    SubOP_decision_list.append(7)
    SubOP_decision_list.append(4)
    SubOP_decision_list.append(3)
    SubOP_decision_list.append(2)

    # SubOPの数値
    SUBOPTION.SubOP_0_fin = ""
    SUBOPTION.SubOP_1_fin = ""
    SUBOPTION.SubOP_2_fin = ""
    SUBOPTION.SubOP_3_fin = ""
    SUBOPTION.SubOP_0_fin_value = ""
    SUBOPTION.SubOP_1_fin_value = ""
    SUBOPTION.SubOP_2_fin_value = ""
    SUBOPTION.SubOP_3_fin_value = ""

    # SubOPのカウント用
    SubOP_counter = []
    SubOP_first_value = []

    # SubOPの初期数決定
    SubOP_BegNum = 0
    SubOP_Random = random.randrange(5)

    if SubOP_Random <= 2:
        SubOP_BegNum = 3
    else:
        SubOP_BegNum = 4

    SUBOPTION.SubOP_fir_NUM = "(初期 "+str(SubOP_BegNum) + "個)"
    SUBOPTION.SubOP_fir_n = SubOP_BegNum

    # SubOPからMainOPを除外

    SubOP_Dec_Num = []
    SubOP_Celect = SubOP_String
    SubOP_Celect_Num = SubOP_Kind_Num
    SubOP_Dec_list = SubOP_decision_list

    if MainOP_NM != 100:
        for i in range(len(eval('SubOP'+str(MainOP_NM)+'_Dec_list'))):
            SubOP_Dec_list.remove(MainOP_NM)

    # SubOPの決定

    # SubOPの決定
    if MainOP_NM != 100:
        for i in range(4):
            rand_a = 0
            rand_a = random.choice(SubOP_Dec_list)
            SubOP_Dec_Num.append(rand_a)
            for n in range(len(eval('SubOP'+str(rand_a)+'_Dec_list'))):
                SubOP_Dec_list.remove(rand_a)
        SubOP_Dec = [SubOP_String[SubOP_Dec_Num[0]], SubOP_String[SubOP_Dec_Num[1]],
                     SubOP_String[SubOP_Dec_Num[2]], SubOP_String[SubOP_Dec_Num[3]]]
        SubOP_Dec_n = SubOP_Dec_Num
    else:
        for i in range(4):
            rand_a = 0
            rand_a = random.choice(SubOP_Dec_list)
            SubOP_Dec_Num.append(rand_a)
            for n in range(len(eval('SubOP'+str(rand_a)+'_Dec_list'))):
                SubOP_Dec_list.remove(rand_a)
        SubOP_Dec = [SubOP_String[SubOP_Dec_Num[0]], SubOP_String[SubOP_Dec_Num[1]],
                     SubOP_String[SubOP_Dec_Num[2]], SubOP_String[SubOP_Dec_Num[3]]]
        SubOP_Dec_n = SubOP_Dec_Num

    for i in range(4):
        SubOP_first_value.append(random.randrange(4)+1)

    # SubOPの強化値
    SubOP0_UP_list = [SubOP_first_value[0]]
    SubOP1_UP_list = [SubOP_first_value[1]]
    SubOP2_UP_list = [SubOP_first_value[2]]
    SubOP3_UP_list = [SubOP_first_value[3]]

    for i in range(4):
        SubOP_pro_list += eval('SubOP'+str(SubOP_Dec_n[i])+'_Dec_list')
    # 強化関数
    for i in range(SubOP_BegNum+1):
        SubOP_UP_kind_rand = random.choice(SubOP_pro_list)
        SubOP_UP_kind = SubOP_Dec_Num.index(SubOP_UP_kind_rand)
        SubOP_UP_value = random.randrange(4)+1
        exec("SubOP" + str(SubOP_UP_kind) +
             "_UP_list.append("+str(SubOP_UP_value)+")")

    # SubOPの計算
    for a in range(4):
        for b in range(len(eval("SubOP"+str(a)+"_UP_list"))):
            exec("SUBOPTION.SubOP_"+str(SubOP_Dec_n[a])+"+=SubOP_"+str(
                SubOP_Dec_n[a])+"_value["+"SubOP"+str(a)+"_UP_list["+str(b)+"]-1]")

    # SubOPカウントの入力
    exec("SUBOPTION.SubOP_"+str(SubOP_Dec_n[0])+"_count = sum(SubOP0_UP_list)")
    exec("SUBOPTION.SubOP_"+str(SubOP_Dec_n[1])+"_count = sum(SubOP1_UP_list)")
    exec("SUBOPTION.SubOP_"+str(SubOP_Dec_n[2])+"_count = sum(SubOP2_UP_list)")
    exec("SUBOPTION.SubOP_"+str(SubOP_Dec_n[3])+"_count = sum(SubOP3_UP_list)")

    SUBOPTION.SubOP_0_fin = SubOP_Dec[0]
    SUBOPTION.SubOP_1_fin = SubOP_Dec[1]
    SUBOPTION.SubOP_2_fin = SubOP_Dec[2]
    SUBOPTION.SubOP_3_fin = SubOP_Dec[3]

    SUBOPTION.SubOP_score_fin = SUBOPTION.SubOP_2 + \
        2*float(format(SUBOPTION.SubOP_7, '.1f')) + SUBOPTION.SubOP_8

    for i in range(4):
        if 2 <= SubOP_Dec_n[i] <= 4 or 6 <= SubOP_Dec_n[i] < 9:
            SUB_TBD = str(
                format(eval("SUBOPTION.SubOP_"+str(SubOP_Dec_n[i])), '.1f')) + "%"
            exec("SUBOPTION.SubOP_"+str(i)+"_fin_value = SUB_TBD")
        else:
            SUB_TBA = format(
                eval("SUBOPTION.SubOP_"+str(SubOP_Dec_n[i])), '0f')
            SUB_TBY = float(SUB_TBA)
            SUB_TBB = format(SUB_TBY, '.0f')
            SUB_TBC = int(SUB_TBB)
            SUB_TBX = '{:,}'.format(SUB_TBC)
            SUB_TBD = str(SUB_TBX)

            exec("SUBOPTION.SubOP_"+str(i)+"_fin_value = SUB_TBD")

# 聖遺物決定部


def Artifact_Simulator(Domain_Number):
    MainOP_N = 0
    Domain_String = "Artifact_name" + str(Domain_Number)  # 秘境番号から文字列作成
    Domain = eval(Domain_String)  # 文字列から配列呼び出し
    # 聖遺物決定
    Artifact_decision = random.randrange(2)
    ART = Domain[Artifact_decision]

    # 部位決定
    Artifact_kind_num = random.randrange(5)
    kind_number_fin = 2*Domain_Number+Artifact_decision
    Kind_String = "Artifact_kind" + str(2*Domain_Number+Artifact_decision)
    Kind_kindname = Artifact_kinds[Artifact_kind_num]
    Kind_list = eval(Kind_String)
    Kind = Kind_list[Artifact_kind_num]

    # MainOP決定
    main = MainOption_object()
    Artifact_MainOP(main, Artifact_kind_num)
    MainOP_N = main.MainOP_Num_fin  # MainOPの種類を記録

    # SubOP決定
    sub = SubOption_object()
    Artifact_SubOP(sub, MainOP_N)

    Main_list = [ART, Kind, main.MainOP_Str_fin, main.MainOP_Val_fin]
    Sub_list = [sub.SubOP_0_fin, sub.SubOP_1_fin, sub.SubOP_2_fin, sub.SubOP_3_fin,
                sub.SubOP_0_fin_value, sub.SubOP_1_fin_value, sub.SubOP_2_fin_value, sub.SubOP_3_fin_value]
    Add_list = [Kind_kindname, sub.SubOP_score_fin,
                sub.SubOP_fir_NUM, Artifact_kind_num, kind_number_fin]
    Sub_OP_Total = [sub.SubOP_0, sub.SubOP_1, sub.SubOP_2, sub.SubOP_3,
                    sub.SubOP_4, sub.SubOP_5, sub.SubOP_6, sub.SubOP_7, sub.SubOP_8, sub.SubOP_9, sub.SubOP_fir_n]

    Artifact_list_fin = Main_list + Sub_list + Add_list+Sub_OP_Total

    return Artifact_list_fin

# 画像作成部


def Create_Artifact_Image(Artifact_kind_name, Artifact_kinds, MO_k, MO_v, SO0_k, SO1_k, SO2_k, SO3_k, SO0_v, SO1_v, SO2_v, SO3_v, Art_n_f, Sub_Score_f, k_p, Kind_n_f):
    img = Image.open('Artifact_Background/Artifact_Background_image.png')
    img_icon = Image.open(
        Artifact_kinds_image[k_p]+"/"+Artifact_kinds_image[k_p]+"_"+str(Kind_n_f)+".png")
    img.paste(img_icon, (600, 130), img_icon)

    x = 75
    y1 = 100
    y0 = 815
    y_a = 20
    y_b = 170

    S0 = "◆ "+SO0_k+"+"+SO0_v
    S1 = "◆ "+SO1_k+"+"+SO1_v
    S2 = "◆ "+SO2_k+"+"+SO2_v
    S3 = "◆ "+SO3_k+"+"+SO3_v

    font1 = ImageFont.truetype('Genshin_font/ja-jp.ttf', 65)
    font2 = ImageFont.truetype('Genshin_font/ja-jp.ttf', 55)
    font3 = ImageFont.truetype('Genshin_font/ja-jp.ttf', 45)
    font4 = ImageFont.truetype('Genshin_font/ja-jp.ttf', 100)
    font5 = ImageFont.truetype('Genshin_font/ja-jp.ttf', 55)
    font6 = ImageFont.truetype('Genshin_font/ja-jp.ttf', 60)
    font7 = ImageFont.truetype('Genshin_font/ja-jp.ttf', 50)
    font8 = ImageFont.truetype('Genshin_font/ja-jp.ttf', 80)
    font9 = ImageFont.truetype('Genshin_font/ja-jp.ttf', 70)

    draw = ImageDraw.Draw(img)
    draw.text((x, 26), Artifact_kind_name, '#FDFBFA', font=font1)
    draw.text((x, 150), Artifact_kinds, '#FDFBFA', font=font2)
    draw.text((x, 360), MO_k, '#BBA79F', font=font3)
    draw.text((x, 415), MO_v, '#FDFBFA', font=font4)
    # SubOP入力
    draw.text((x+5, y0), S0, '#495366', font=font5)
    draw.text((x+5, y0+y1), S1, '#495366', font=font5)
    draw.text((x+5, y0+y1*2), S2, '#495366', font=font5)
    draw.text((x+5, y0+y1*3), S3, '#495366', font=font5)
    # 聖遺物名
    draw.text((x+60, y0+y1*4+27+y_a), "聖遺物効果 :", '#60B35A', font=font7)
    draw.text((x, y0+y1*4+20+y_a), "                     " +
              Art_n_f, '#60B35A', font=font6)
    draw.text((x-30, y0+y1*4+27+y_b+5), "◇ スコア :", '#495366', font=font9)
    draw.text((x+750, y0+y1*4+27+y_b), str(Sub_Score_f), '#495366', font=font8)

    # 画像保存(ファイルは無し)
    fileio = io.BytesIO()
    img.save(fileio, format="png")
    fileio.seek(0)
    return fileio


# Excel作成部(テスト)
def Create_Artifact_Excel(l_f, h, T_n):

    wb = openpyxl.Workbook()
    ws = wb['Sheet']
    ws.title = '聖遺物厳選結果'

    length = len(l_f)

    # 幅調整(A~R)
    ws.column_dimensions['A'].width = 2.07*4.5  # 0.番号
    ws.column_dimensions['B'].width = 2.07*10   # 1.聖遺物セット名
    ws.column_dimensions['C'].width = 7.2       # 2.部位
    ws.column_dimensions['D'].width = 2.07*10   # 3.聖遺物名
    ws.column_dimensions['E'].width = 19.4        # 4.MainOPの種類
    ws.column_dimensions['F'].width = 9.4       # 5.MainOPの数値
    ws.column_dimensions['G'].width = 8.2       # 6.Score
    ws.column_dimensions['H'].width = 8.2       # 7.初期数
    ws.column_dimensions['I'].width = 11.3  # 8.HP実数
    ws.column_dimensions['J'].width = 11.3  # 9.攻撃力実数
    ws.column_dimensions['K'].width = 11.3   # 10.防御力実数
    ws.column_dimensions['L'].width = 9.0  # 11.HP%
    ws.column_dimensions['M'].width = 9.0  # 12.攻撃力%
    ws.column_dimensions['N'].width = 9.0  # 13.防御力%
    ws.column_dimensions['O'].width = 9.9  # 14.会心率
    ws.column_dimensions['P'].width = 9.9  # 15.会心ダメ
    ws.column_dimensions['Q'].width = 9.9  # 16.元素熟知
    ws.column_dimensions['R'].width = 11  # 17.チャージ効率

    # 高さの調整(1~試行回数+1)
    ws.row_dimensions[1].height = 22.20   # headerの高さ
    for i in range(T_n):
        ws.row_dimensions[i+2].height = 18   # それ以外の高さ

    # 数値入力
    ws.append(h)
    for i in range(len(l_f)):
        ws.append(l_f[i])

    # 中央揃え
    for row in ws["A:R"]:
        for cell in row:
            cell.alignment = Alignment(
                horizontal="centerContinuous", vertical="center")

    # Tableの設定
    table = Table(displayName='Table1', ref='A1:R'+str(T_n+1))
    table_style = TableStyleInfo(
        name='TableStyleLight9')

    # テーブルのスタイルを設定
    table.tableStyleInfo = table_style
    ws.add_table(table)

    # 罫線
    # 種類
    side1 = Side(style='medium', color='000000')  # 極太罫線
    side2 = Side(style='double', color='000000')  # 二重罫線
    # 罫線の場所

    # 上下の罫線
    Border_UP_0 = Border(top=side1, bottom=side2)
    Border_UP_1 = Border(top=side1, bottom=side2, left=side1, right=side1)
    Border_UP_2 = Border(top=side1, bottom=side2, left=side1)
    Border_UP_3 = Border(top=side1, bottom=side2, right=side1)

    Border_UP2_0 = Border(top=side2)
    Border_UP2_1 = Border(top=side2, left=side1, right=side1)
    Border_UP2_2 = Border(top=side2, left=side1)
    Border_UP2_3 = Border(top=side2, right=side1)

    Border_MID_0 = Border()
    Border_MID_1 = Border(left=side1, right=side1)
    Border_MID_2 = Border(left=side1)
    Border_MID_3 = Border(right=side1)

    Border_BOT_0 = Border(bottom=side1)
    Border_BOT_1 = Border(bottom=side1, left=side1, right=side1)
    Border_BOT_2 = Border(bottom=side1, left=side1)
    Border_BOT_3 = Border(bottom=side1, right=side1)

    # 縦の罫線(1 : 1,2   2 : 3,5,7,9  3 : 4,6,8,18)
    for i in range(T_n+1):
        if i == 0:
            for n in range(18):
                if n == 0 or n == 1:
                    ws.cell(row=i+1, column=n+1).border = Border_UP_1
                elif n == 2 or n == 4 or n == 6 or n == 8:
                    ws.cell(row=i+1, column=n+1).border = Border_UP_2
                elif n == 3 or n == 5 or n == 7 or n == 17:
                    ws.cell(row=i+1, column=n+1).border = Border_UP_3
                else:
                    ws.cell(row=i+1, column=n+1).border = Border_UP_0
        elif i == 1:
            for n in range(18):
                if n == 0 or n == 1:
                    ws.cell(row=i+1, column=n+1).border = Border_UP2_1
                elif n == 2 or n == 4 or n == 6 or n == 8:
                    ws.cell(row=i+1, column=n+1).border = Border_UP2_2
                elif n == 3 or n == 5 or n == 7 or n == 17:
                    ws.cell(row=i+1, column=n+1).border = Border_UP2_3
                else:
                    ws.cell(row=i+1, column=n+1).border = Border_UP2_0
        elif i == T_n:
            for n in range(18):
                if n == 0 or n == 1:
                    ws.cell(row=i+1, column=n+1).border = Border_BOT_1
                elif n == 2 or n == 4 or n == 6 or n == 8:
                    ws.cell(row=i+1, column=n+1).border = Border_BOT_2
                elif n == 3 or n == 5 or n == 7 or n == 17:
                    ws.cell(row=i+1, column=n+1).border = Border_BOT_3
                else:
                    ws.cell(row=i+1, column=n+1).border = Border_BOT_0
        else:
            for n in range(18):
                if n == 0 or n == 1:
                    ws.cell(row=i+1, column=n+1).border = Border_MID_1
                elif n == 2 or n == 4 or n == 6 or n == 8:
                    ws.cell(row=i+1, column=n+1).border = Border_MID_2
                elif n == 3 or n == 5 or n == 7 or n == 17:
                    ws.cell(row=i+1, column=n+1).border = Border_MID_3
                else:
                    ws.cell(row=i+1, column=n+1).border = Border_MID_0

    # 色の塗り変え
    fill = openpyxl.styles.PatternFill(
        patternType='solid', fgColor='548235', bgColor='548235')
    ws['I1'].fill = fill
    ws['J1'].fill = fill
    ws['K1'].fill = fill
    ws['L1'].fill = fill
    ws['M1'].fill = fill
    ws['N1'].fill = fill
    ws['O1'].fill = fill
    ws['P1'].fill = fill
    ws['Q1'].fill = fill
    ws['R1'].fill = fill

    fileio2 = io.BytesIO()
    wb.save(fileio2)
    fileio2.seek(0)

    return fileio2


def excel_create(l_f, h):

    wb = openpyxl.Workbook()
    ws = wb['Sheet']
    ws.title = '聖遺物リスト'
    # 幅,高さ調整
    ws.column_dimensions['A'].width = 2.07*2
    for i in range(ord('B'), ord('B')+len(h)+2):
        ws.column_dimensions[chr(i)].width = 2.07*10
    ws.row_dimensions[1].height = 20
    for i in range(2, len(l_f)+2):
        ws.row_dimensions[i].height = 15

    # 数値入力
    ws.append(h)
    for i in range(len(l_f)):
        ws.append(l_f[i])

    # 中央揃え
    for row in ws["A:"+chr(ord('A')+len(h)-1)]:
        for cell in row:
            cell.alignment = Alignment(
                horizontal="centerContinuous", vertical="center")

    table = Table(displayName='Table1', ref='A1:' +
                  chr(ord('A')+len(h)-1)+str(len(l_f)+1))
    table_style = TableStyleInfo(
        name='TableStyleMedium1')

    # テーブルのスタイルを設定
    table.tableStyleInfo = table_style
    ws.add_table(table)

    fileio1 = io.BytesIO()
    wb.save(fileio1)
    fileio1.seek(0)

    return fileio1


# テスト用(聖遺物の名前をExcelに出力)
list_fin = []
for i in range(13):
    list1 = []
    list2 = []

    list1.append((i*2))
    list2.append((i*2+1))

    list1.append(eval('Artifact_name'+str(i)+'[0]'))
    list2.append(eval('Artifact_name'+str(i)+'[1]'))

    for n in range(5):
        list1.append(eval('Artifact_kind'+str(2*i)+'['+str(n)+']'))
        list2.append(eval('Artifact_kind'+str(2*i+1)+'['+str(n)+']'))

    list_fin.append(list1)
    list_fin.append(list2)


# ログイン


@ client.event
async def on_ready():
    print(f'We have logged in as {client.user}')

# 入力時の反応


@ client.event
async def on_message(message):
    if message.author == client.user:
        return

    # テキスト出力
    # 0. '剣闘士のフィナーレ', '大地を流浪する楽団'
    if message.content.startswith('/Artifact00'):
        x = Artifact_Simulator(0)
        # Embed
        embed_Resule_fin = discord.Embed(title='◆ 出力結果', description="◆ 聖遺物名："+x[0]+"\n◆ " + x[12]+": "+x[1]+"\n◆ MainOP："+x[2]+'  '+x[3]+"\n◆ SubOP：" +
                                         x[14]+"\n・ "+x[4]+" : "+x[8]+"\n・ "+x[5]+" : "+x[9]+"\n・ "+x[6]+" : "+x[10]+"\n・ "+x[7]+" : "+x[11]+"\n☆ スコア："+str(format(x[13], '.1f')))
        await message.channel.send(embed=embed_Resule_fin)
    # 1. '雷のような怒り', '雷を鎮める尊者'
    if message.content.startswith('/Artifact01'):
        x = Artifact_Simulator(1)
        # Embed
        embed_Resule_fin = discord.Embed(title='◆ 出力結果', description="◆ 聖遺物名："+x[0]+"\n◆ " + x[12]+": "+x[1]+"\n◆ MainOP："+x[2]+'  '+x[3]+"\n◆ SubOP：" +
                                         x[14]+"\n・ "+x[4]+" : "+x[8]+"\n・ "+x[5]+" : "+x[9]+"\n・ "+x[6]+" : "+x[10]+"\n・ "+x[7]+" : "+x[11]+"\n☆ スコア："+str(format(x[13], '.1f')))
        await message.channel.send(embed=embed_Resule_fin)
    # 2. '翠緑の影', '愛される少女'
    if message.content.startswith('/Artifact02'):
        x = Artifact_Simulator(2)
        # Embed
        embed_Resule_fin = discord.Embed(title='◆ 出力結果', description="◆ 聖遺物名："+x[0]+"\n◆ " + x[12]+": "+x[1]+"\n◆ MainOP："+x[2]+'  '+x[3]+"\n◆ SubOP：" +
                                         x[14]+"\n・ "+x[4]+" : "+x[8]+"\n・ "+x[5]+" : "+x[9]+"\n・ "+x[6]+" : "+x[10]+"\n・ "+x[7]+" : "+x[11]+"\n☆ スコア："+str(format(x[13], '.1f')))
        await message.channel.send(embed=embed_Resule_fin)
    # 3. '氷風を彷徨う勇士', '沈淪の心'
    if message.content.startswith('/Artifact03'):
        x = Artifact_Simulator(3)
        # Embed
        embed_Resule_fin = discord.Embed(title='◆ 出力結果', description="◆ 聖遺物名："+x[0]+"\n◆ " + x[12]+": "+x[1]+"\n◆ MainOP："+x[2]+'  '+x[3]+"\n◆ SubOP：" +
                                         x[14]+"\n・ "+x[4]+" : "+x[8]+"\n・ "+x[5]+" : "+x[9]+"\n・ "+x[6]+" : "+x[10]+"\n・ "+x[7]+" : "+x[11]+"\n☆ スコア："+str(format(x[13], '.1f')))
        await message.channel.send(embed=embed_Resule_fin)
    # 4. '燃え盛る炎の魔女', '烈火を渡る賢者'
    if message.content.startswith('/Artifact04'):
        x = Artifact_Simulator(4)
        # Embed
        embed_Resule_fin = discord.Embed(title='◆ 出力結果', description="◆ 聖遺物名："+x[0]+"\n◆ " + x[12]+": "+x[1]+"\n◆ MainOP："+x[2]+'  '+x[3]+"\n◆ SubOP：" +
                                         x[14]+"\n・ "+x[4]+" : "+x[8]+"\n・ "+x[5]+" : "+x[9]+"\n・ "+x[6]+" : "+x[10]+"\n・ "+x[7]+" : "+x[11]+"\n☆ スコア："+str(format(x[13], '.1f')))
        await message.channel.send(embed=embed_Resule_fin)
    # 5. '旧貴族のしつけ', '血染めの騎士道'
    if message.content.startswith('/Artifact05'):
        x = Artifact_Simulator(5)
        # Embed
        embed_Resule_fin = discord.Embed(title='◆ 出力結果', description="◆ 聖遺物名："+x[0]+"\n◆ " + x[12]+": "+x[1]+"\n◆ MainOP："+x[2]+'  '+x[3]+"\n◆ SubOP：" +
                                         x[14]+"\n・ "+x[4]+" : "+x[8]+"\n・ "+x[5]+" : "+x[9]+"\n・ "+x[6]+" : "+x[10]+"\n・ "+x[7]+" : "+x[11]+"\n☆ スコア："+str(format(x[13], '.1f')))
        await message.channel.send(embed=embed_Resule_fin)
    # 6. '悠久の磐岩', '逆飛びの流星'
    if message.content.startswith('/Artifact06'):
        x = Artifact_Simulator(6)
        # Embed
        embed_Resule_fin = discord.Embed(title='◆ 出力結果', description="◆ 聖遺物名："+x[0]+"\n◆ " + x[12]+": "+x[1]+"\n◆ MainOP："+x[2]+'  '+x[3]+"\n◆ SubOP：" +
                                         x[14]+"\n・ "+x[4]+" : "+x[8]+"\n・ "+x[5]+" : "+x[9]+"\n・ "+x[6]+" : "+x[10]+"\n・ "+x[7]+" : "+x[11]+"\n☆ スコア："+str(format(x[13], '.1f')))
        await message.channel.send(embed=embed_Resule_fin)
    # 7. '千岩牢固', '蒼白の炎'
    if message.content.startswith('/Artifact07'):
        x = Artifact_Simulator(7)
        # Embed
        embed_Resule_fin = discord.Embed(title='◆ 出力結果', description="◆ 聖遺物名："+x[0]+"\n◆ " + x[12]+": "+x[1]+"\n◆ MainOP："+x[2]+'  '+x[3]+"\n◆ SubOP：" +
                                         x[14]+"\n・ "+x[4]+" : "+x[8]+"\n・ "+x[5]+" : "+x[9]+"\n・ "+x[6]+" : "+x[10]+"\n・ "+x[7]+" : "+x[11]+"\n☆ スコア："+str(format(x[13], '.1f')))
        await message.channel.send(embed=embed_Resule_fin)
    # 8. '辰砂往生録', '来歆の余響'
    if message.content.startswith('/Artifact08'):
        x = Artifact_Simulator(8)
        # Embed
        embed_Resule_fin = discord.Embed(title='◆ 出力結果', description="◆ 聖遺物名："+x[0]+"\n◆ " + x[12]+": "+x[1]+"\n◆ MainOP："+x[2]+'  '+x[3]+"\n◆ SubOP：" +
                                         x[14]+"\n・ "+x[4]+" : "+x[8]+"\n・ "+x[5]+" : "+x[9]+"\n・ "+x[6]+" : "+x[10]+"\n・ "+x[7]+" : "+x[11]+"\n☆ スコア："+str(format(x[13], '.1f')))
        await message.channel.send(embed=embed_Resule_fin)
    # 9. '追憶のしめ縄', '絶縁の旗印'
    if message.content.startswith('/Artifact09'):
        x = Artifact_Simulator(9)
        # Embed
        embed_Resule_fin = discord.Embed(title='◆ 出力結果', description="◆ 聖遺物名："+x[0]+"\n◆ " + x[12]+": "+x[1]+"\n◆ MainOP："+x[2]+'  '+x[3]+"\n◆ SubOP：" +
                                         x[14]+"\n・ "+x[4]+" : "+x[8]+"\n・ "+x[5]+" : "+x[9]+"\n・ "+x[6]+" : "+x[10]+"\n・ "+x[7]+" : "+x[11]+"\n☆ スコア："+str(format(x[13], '.1f')))
        await message.channel.send(embed=embed_Resule_fin)
    # 10. '華館夢醒形骸記', '海染硨磲'
    if message.content.startswith('/Artifact10'):
        x = Artifact_Simulator(10)
        # Embed
        embed_Resule_fin = discord.Embed(title='◆ 出力結果', description="◆ 聖遺物名："+x[0]+"\n◆ " + x[12]+": "+x[1]+"\n◆ MainOP："+x[2]+'  '+x[3]+"\n◆ SubOP：" +
                                         x[14]+"\n・ "+x[4]+" : "+x[8]+"\n・ "+x[5]+" : "+x[9]+"\n・ "+x[6]+" : "+x[10]+"\n・ "+x[7]+" : "+x[11]+"\n☆ スコア："+str(format(x[13], '.1f')))
        await message.channel.send(embed=embed_Resule_fin)
    # 11. '森林の記憶', '金メッキの夢'
    if message.content.startswith('/Artifact11'):
        x = Artifact_Simulator(11)
        # Embed
        embed_Resule_fin = discord.Embed(title='◆ 出力結果', description="◆ 聖遺物名："+x[0]+"\n◆ " + x[12]+": "+x[1]+"\n◆ MainOP："+x[2]+'  '+x[3]+"\n◆ SubOP：" +
                                         x[14]+"\n・ "+x[4]+" : "+x[8]+"\n・ "+x[5]+" : "+x[9]+"\n・ "+x[6]+" : "+x[10]+"\n・ "+x[7]+" : "+x[11]+"\n☆ スコア："+str(format(x[13], '.1f')))
        await message.channel.send(embed=embed_Resule_fin)
    # 12. '砂上の楼閣の史話', '楽園の絶花'
    if message.content.startswith('/Artifact12'):
        x = Artifact_Simulator(12)
        # Embed
        embed_Resule_fin = discord.Embed(title='◆ 出力結果', description="◆ 聖遺物名："+x[0]+"\n◆ " + x[12]+": "+x[1]+"\n◆ MainOP："+x[2]+'  '+x[3]+"\n◆ SubOP：" +
                                         x[14]+"\n・ "+x[4]+" : "+x[8]+"\n・ "+x[5]+" : "+x[9]+"\n・ "+x[6]+" : "+x[10]+"\n・ "+x[7]+" : "+x[11]+"\n☆ スコア："+str(format(x[13], '.1f')))
        await message.channel.send(embed=embed_Resule_fin)

    # 画像出力
    # 0. '剣闘士のフィナーレ', '大地を流浪する楽団'
    if message.content.startswith('/img_Artifact00'):
        x = Artifact_Simulator(0)
        files = Create_Artifact_Image(x[1], x[12], x[2], x[3], x[4], x[5], x[6],
                                      x[7], x[8], x[9], x[10], x[11], x[0], format(x[13], '.1f'), x[15], x[16])
        # embed設定
        Article_result_img = discord.File(files, filename="Article_img.png")
        embed_Resule_fin = discord.Embed(title="◆ 出力結果 : "+x[14])
        embed_Resule_fin.set_image(url="attachment://Article_img.png")
        await message.channel.send(file=Article_result_img, embed=embed_Resule_fin)
    # 1. '雷のような怒り', '雷を鎮める尊者'
    if message.content.startswith('/img_Artifact01'):
        x = Artifact_Simulator(1)
        files = Create_Artifact_Image(x[1], x[12], x[2], x[3], x[4], x[5], x[6],
                                      x[7], x[8], x[9], x[10], x[11], x[0], format(x[13], '.1f'), x[15], x[16])

        # embed設定
        Article_result_img = discord.File(files, filename="Article_img.png")
        embed_Resule_fin = discord.Embed(title="◆ 出力結果 : "+x[14])
        embed_Resule_fin.set_image(url="attachment://Article_img.png")
        await message.channel.send(file=Article_result_img, embed=embed_Resule_fin)
    # 2. '翠緑の影', '愛される少女'
    if message.content.startswith('/img_Artifact02'):
        x = Artifact_Simulator(2)
        files = Create_Artifact_Image(x[1], x[12], x[2], x[3], x[4], x[5], x[6],
                                      x[7], x[8], x[9], x[10], x[11], x[0], format(x[13], '.1f'), x[15], x[16])

        # embed設定
        Article_result_img = discord.File(files, filename="Article_img.png")
        embed_Resule_fin = discord.Embed(title="◆ 出力結果 : "+x[14])
        embed_Resule_fin.set_image(url="attachment://Article_img.png")
        await message.channel.send(file=Article_result_img, embed=embed_Resule_fin)
    # 3. '氷風を彷徨う勇士', '沈淪の心'
    if message.content.startswith('/img_Artifact03'):
        x = Artifact_Simulator(3)
        files = Create_Artifact_Image(x[1], x[12], x[2], x[3], x[4], x[5], x[6],
                                      x[7], x[8], x[9], x[10], x[11], x[0], format(x[13], '.1f'), x[15], x[16])

        # embed設定
        Article_result_img = discord.File(files, filename="Article_img.png")
        embed_Resule_fin = discord.Embed(title="◆ 出力結果 : "+x[14])
        embed_Resule_fin.set_image(url="attachment://Article_img.png")
        await message.channel.send(file=Article_result_img, embed=embed_Resule_fin)
    # 4. '燃え盛る炎の魔女', '烈火を渡る賢者'
    if message.content.startswith('/img_Artifact04'):
        x = Artifact_Simulator(4)
        files = Create_Artifact_Image(x[1], x[12], x[2], x[3], x[4], x[5], x[6],
                                      x[7], x[8], x[9], x[10], x[11], x[0], format(x[13], '.1f'), x[15], x[16])

        # embed設定
        Article_result_img = discord.File(files, filename="Article_img.png")
        embed_Resule_fin = discord.Embed(title="◆ 出力結果 : "+x[14])
        embed_Resule_fin.set_image(url="attachment://Article_img.png")
        await message.channel.send(file=Article_result_img, embed=embed_Resule_fin)
    # 5. '旧貴族のしつけ', '血染めの騎士道'
    if message.content.startswith('/img_Artifact05'):
        x = Artifact_Simulator(5)
        files = Create_Artifact_Image(x[1], x[12], x[2], x[3], x[4], x[5], x[6],
                                      x[7], x[8], x[9], x[10], x[11], x[0], format(x[13], '.1f'), x[15], x[16])

        # embed設定
        Article_result_img = discord.File(files, filename="Article_img.png")
        embed_Resule_fin = discord.Embed(title="◆ 出力結果 : "+x[14])
        embed_Resule_fin.set_image(url="attachment://Article_img.png")
        await message.channel.send(file=Article_result_img, embed=embed_Resule_fin)
    # 6. '悠久の磐岩', '逆飛びの流星'
    if message.content.startswith('/img_Artifact06'):
        x = Artifact_Simulator(6)
        files = Create_Artifact_Image(x[1], x[12], x[2], x[3], x[4], x[5], x[6],
                                      x[7], x[8], x[9], x[10], x[11], x[0], format(x[13], '.1f'), x[15], x[16])

        # embed設定
        Article_result_img = discord.File(files, filename="Article_img.png")
        embed_Resule_fin = discord.Embed(title="◆ 出力結果 : "+x[14])
        embed_Resule_fin.set_image(url="attachment://Article_img.png")
        await message.channel.send(file=Article_result_img, embed=embed_Resule_fin)
    # 7. '千岩牢固', '蒼白の炎'
    if message.content.startswith('/img_Artifact07'):
        x = Artifact_Simulator(7)
        files = Create_Artifact_Image(x[1], x[12], x[2], x[3], x[4], x[5], x[6],
                                      x[7], x[8], x[9], x[10], x[11], x[0], format(x[13], '.1f'), x[15], x[16])

        # embed設定
        Article_result_img = discord.File(files, filename="Article_img.png")
        embed_Resule_fin = discord.Embed(title="◆ 出力結果 : "+x[14])
        embed_Resule_fin.set_image(url="attachment://Article_img.png")
        await message.channel.send(file=Article_result_img, embed=embed_Resule_fin)
    # 8. '辰砂往生録', '来歆の余響'
    if message.content.startswith('/img_Artifact08'):
        x = Artifact_Simulator(8)
        files = Create_Artifact_Image(x[1], x[12], x[2], x[3], x[4], x[5], x[6],
                                      x[7], x[8], x[9], x[10], x[11], x[0], format(x[13], '.1f'), x[15], x[16])

        # embed設定
        Article_result_img = discord.File(files, filename="Article_img.png")
        embed_Resule_fin = discord.Embed(title="◆ 出力結果 : "+x[14])
        embed_Resule_fin.set_image(url="attachment://Article_img.png")
        await message.channel.send(file=Article_result_img, embed=embed_Resule_fin)
    # 9. '追憶のしめ縄', '絶縁の旗印'
    if message.content.startswith('/img_Artifact09'):
        x = Artifact_Simulator(9)
        files = Create_Artifact_Image(x[1], x[12], x[2], x[3], x[4], x[5], x[6],
                                      x[7], x[8], x[9], x[10], x[11], x[0], format(x[13], '.1f'), x[15], x[16])

        # embed設定
        Article_result_img = discord.File(files, filename="Article_img.png")
        embed_Resule_fin = discord.Embed(title="◆ 出力結果 : "+x[14])
        embed_Resule_fin.set_image(url="attachment://Article_img.png")
        await message.channel.send(file=Article_result_img, embed=embed_Resule_fin)
    # 10. '華館夢醒形骸記', '海染硨磲'
    if message.content.startswith('/img_Artifact10'):
        x = Artifact_Simulator(10)
        files = Create_Artifact_Image(x[1], x[12], x[2], x[3], x[4], x[5], x[6],
                                      x[7], x[8], x[9], x[10], x[11], x[0], format(x[13], '.1f'), x[15], x[16])

        # embed設定
        Article_result_img = discord.File(files, filename="Article_img.png")
        embed_Resule_fin = discord.Embed(title="◆ 出力結果 : "+x[14])
        embed_Resule_fin.set_image(url="attachment://Article_img.png")
        await message.channel.send(file=Article_result_img, embed=embed_Resule_fin)
    # 11. '森林の記憶', '金メッキの夢'
    if message.content.startswith('/img_Artifact11'):
        x = Artifact_Simulator(11)
        files = Create_Artifact_Image(x[1], x[12], x[2], x[3], x[4], x[5], x[6],
                                      x[7], x[8], x[9], x[10], x[11], x[0], format(x[13], '.1f'), x[15], x[16])

        await message.channel.send(x[14], file=discord.File(files, "Article_img.png"))
    # 12. '砂上の楼閣の史話', '楽園の絶花'
    if message.content.startswith('/img_Artifact12'):
        x = Artifact_Simulator(12)
        files = Create_Artifact_Image(x[1], x[12], x[2], x[3], x[4], x[5], x[6],
                                      x[7], x[8], x[9], x[10], x[11], x[0], format(x[13], '.1f'), x[15], x[16])

        # embed設定
        Article_result_img = discord.File(files, filename="Article_img.png")
        embed_Resule_fin = discord.Embed(title="◆ 出力結果 : "+x[14])
        embed_Resule_fin.set_image(url="attachment://Article_img.png")
        await message.channel.send(file=Article_result_img, embed=embed_Resule_fin)

    # エクセル出力
    if message.content.startswith('/xlsx_Artifact'):
        art_d = Artifact_xlsx_Dec()
        Art_n = 0
        Tri_n = 0
        Artifact_AAA = ''
        Trial_AAA = ''
        Article_list_fin = []
        image_file = discord.File(
            "Articlename_table.png", filename="Articlename_table.png")
        icon_first = discord.File(
            "Flower/Flower_0.png", filename="Flower_0.png")
        channel = message.channel

        # embedの設定
        embed_Artifact_num = discord.Embed(
            title="◇ 聖遺物秘境選択", description="※ 厳選したい聖遺物秘境番号を入力してください(数値のみ)")
        embed_Artifact_num.set_author(
            name="複数回試行(Excelファイル出力)")
        embed_Artifact_num.set_image(url="attachment://Articlename_table.png")
        embed_Trial_num = discord.Embed(
            title="◇ 試行回数", description="※ 1~1500の間の数値で入力してください")
        embed_Trial_num.set_author(
            name="複数回試行(Excelファイル出力)")
        embed_Error = discord.Embed(
            title="Error", description="※ 入力値が不正です。再度試行してください")
        embed_Error.set_author(
            name="複数回試行(Excelファイル出力)")
        embed_Result = discord.Embed(
            title="◇ 試行中", description="※ 数秒後に出力されます。")
        embed_Result.set_author(
            name="複数回試行(Excelファイル出力)")

        await channel.send(file=image_file, embed=embed_Artifact_num)

        def check_Artifact(m):
            art_d.Artifact_Number_0 = m.content
            return m.channel == channel and m.content
        try:
            Artifact_AAA = await client.wait_for('message', check=check_Artifact, timeout=20)
        except asyncio.TimeoutError:
            await channel.send('時間切れです')
        else:
            if art_d.Artifact_Number_0.isdecimal():
                Art_n = int(art_d.Artifact_Number_0)
                if Art_n <= 12:
                    await channel.send(embed=embed_Trial_num)

                    def check_Trial(m):
                        art_d.Trial_Number_0 = m.content
                        return m.channel == channel and m.content
                    try:
                        Trial_AAA = await client.wait_for('message', check=check_Trial, timeout=20)
                    except asyncio.TimeoutError:
                        await channel.send('時間切れです')
                    else:
                        if art_d.Trial_Number_0.isdecimal():
                            Tri_n = int(art_d.Trial_Number_0)
                            if 1 <= Tri_n <= 1500:
                                await message.channel.send(embed=embed_Result)
                                for i in range(Tri_n):
                                    x = Artifact_Simulator(Art_n)
                                    Article_list_fin.append([i+1, x[0], Artifact_kinds[x[15]], x[1], x[2], x[3], x[13],
                                                            x[27], x[17], x[18], x[26], x[21], x[19], x[20], x[24], x[25], x[22], x[23]])
                                files = Create_Artifact_Excel(
                                    Article_list_fin, header_Article, Tri_n)
                                await message.channel.send(file=discord.File(files, "Article_Trial.xlsx"))
                            else:
                                await channel.send(embed=embed_Error)
                        else:
                            await channel.send(embed=embed_Error)
                else:
                    await channel.send(embed=embed_Error)
            else:
                await channel.send(embed=embed_Error)

    # 聖遺物のセット表を出力
    if message.content.startswith('/xlsx_Test'):
        files = excel_create(list_fin, header)
        await message.channel.send(file=discord.File(files, "file.xlsx"))

    # 聖遺物のセット表を出力
    if message.content.startswith('/Test_xlsx'):

        Artifact_num = 0
        Trial_num = 2000
        Article_list_fin = []
        t1 = time.time()

        for i in range(Trial_num):
            x = Artifact_Simulator(Artifact_num)
            print(x[27])
            Article_list_fin.append([i+1, x[0], Artifact_kinds[x[15]], x[1], x[2], x[3], x[13], x[27],
                                    x[17], x[18], x[26], x[21], x[19], x[20], x[24], x[25], x[22], x[23]])
        t2 = time.time()

        files = Create_Artifact_Excel(
            Article_list_fin, header_Article, Trial_num)
        t3 = time.time()
        print("所要時間1："+str(t2-t1))
        print("所要時間2："+str(t3-t2))
        await message.channel.send(file=discord.File(files, "Article_Trial.xlsx"))

    if message.content.startswith('/help'):
        image_file = discord.File(
            "Articlename_table.png", filename="Articlename_table.png")
        icon_flower = discord.File(
            "Flower/Flower_0.png", filename="Flower_0.png")
        embed_help = discord.Embed(title='◆ ヘルプ')
        embed_help_twitter = discord.Embed(
            title="Twitterアカウントはこちらから",
            url="https://twitter.com/Art_Gri_Sim_121"
        )
        embed_help_twitter.set_author(
            name="【原神】聖遺物厳選シミュレータ discord bot",
            url="https://twitter.com/Art_Gri_Sim_121"
        )
        embed_help_twitter.set_image(url='attachment://Flower_0.png')

        embed_help.add_field(
            name="①  「/Artifact00」", value="・ 1つの聖遺物結果をテキストで出力します", inline=False)
        embed_help.add_field(
            name="②  「/img_Artifact00」", value="・ 1つの聖遺物結果を画像で出力します", inline=False)
        embed_help.add_field(name="",
                             value="※ 00~12 に変えることで聖遺物秘境を選択できます", inline=False)

        embed_help.add_field(
            name="③  「/xlsx_Artifact」", value="・ 複数の聖遺物厳選の結果をエクセルファイルにて出力します", inline=False)
        embed_help.add_field(
            name="", value='※1 コマンドを実行すると入力が求められます', inline=False)
        embed_help.add_field(
            name="", value='※2 聖遺物番号は1~12で入力してください', inline=False)

        embed_help.add_field(
            name="◇ 聖遺物秘境番号", value="・ 以下の表を見て選択してください", inline=False)
        embed_help.set_image(url='attachment://Articlename_table.png')
        await message.channel.send(file=image_file, embed=embed_help)
        await message.channel.send(file=icon_flower, embed=embed_help_twitter)

client.run(TOKEN)


# header_Article_Main = ['No.', '聖遺物', '部位', '聖遺物名','Main Op', '値', 'Score', '初期数']
# header_Article_Sub = ['HP実数','攻撃力実数','防御力実数','HP%','攻撃力%','防御力%','会心率','会心ダメ','元素熟知','チャージ効率']


# x[0] : 聖遺物セット名
# x[1] : 聖遺物名
# x[2] : MainOP
# x[3] : MainOPの値
# x[4] : SubOP0
# x[5] : SubOP1
# x[6] : SubOP2
# x[7] : SubOP3
# x[8] : SubOP0の値
# x[9] : SubOP1の値
# x[10] : SubOP2の値
# x[11] : SubOP3の値
# x[12] : 聖遺物の部位
# x[13] : スコア(int)
# x[14] : SubOPの初期数
# x[15] : 聖遺物の部位
# x[16] : 聖遺物セットの種類(数値)
# x[17] : HP実数
# x[18] : 攻撃力実数
# x[19] : 攻撃力%
# x[20] : 防御力%
# x[21] : HP%
# x[22] : 元素熟知
# x[23] : 元素チャージ効率
# x[24] : 会心率
# x[25] : 会心ダメ
# x[26] : 防御力実数
# x[27] : 初期数(数値)
